from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, Response
from lark_client import LarkClient
from database import (
    init_db, insert_employees, clear_employees, query_employees, get_employee_stats,
    insert_records, clear_records, clear_records_by_date, query_records, get_stats, get_db,
)
import threading

app = Flask(__name__)
app.secret_key = "lark-employee-record-secret"

init_db()
lark = LarkClient()

LARK_BASE_APP_TOKEN = "GNjTbTciAaqrFcs5YoLl8kNCgAd"
LARK_BASE_TABLE_ID = "tblj8Q7JAhCN3Y0g"


def _extract_text(val):
    if isinstance(val, list) and val:
        return val[0].get("text", "")
    elif isinstance(val, dict):
        return val.get("text", "")
    return str(val) if val else ""


def _push_to_lark_base(records, replace_all=True, append_only=False, upsert=False):
    import requests as req
    try:
        token = lark._get_tenant_access_token()
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        base_url = f"https://open.larksuite.com/open-apis/bitable/v1/apps/{LARK_BASE_APP_TOKEN}/tables/{LARK_BASE_TABLE_ID}"

        if append_only or upsert:
            existing_map = {}
            page_token = None
            while True:
                params = {"page_size": 500}
                if page_token:
                    params["page_token"] = page_token
                existing = req.post(f"{base_url}/records/search", headers=headers, json=params)
                existing_data = existing.json()
                items = existing_data.get("data", {}).get("items", [])
                for item in items:
                    f = item.get("fields", {})
                    key = f"{_extract_text(f.get('ID', ''))}|{_extract_text(f.get('Date', ''))}"
                    existing_map[key] = item["record_id"]
                if not existing_data.get("data", {}).get("has_more"):
                    break
                page_token = existing_data.get("data", {}).get("page_token")

            if append_only:
                new_records = [r for r in records if f"{r['employee_id']}|{r['work_date']}" not in existing_map]
                if not new_records:
                    return len(records), 0
                records = new_records

        if replace_all and not append_only and not upsert:
            all_ids = list(existing_map.values()) if upsert else []
            if not all_ids:
                existing_ids = []
                page_token = None
                while True:
                    params = {"page_size": 500}
                    if page_token:
                        params["page_token"] = page_token
                    existing = req.post(f"{base_url}/records/search", headers=headers, json=params)
                    existing_data = existing.json()
                    items = existing_data.get("data", {}).get("items", [])
                    if items:
                        existing_ids.extend([item["record_id"] for item in items])
                    if not existing_data.get("data", {}).get("has_more"):
                        break
                    page_token = existing_data.get("data", {}).get("page_token")
                all_ids = existing_ids

            if all_ids:
                for i in range(0, len(all_ids), 500):
                    batch = all_ids[i:i+500]
                    req.post(f"{base_url}/records/batch_delete", headers=headers, json={"records": batch})

        pushed = 0
        if upsert:
            to_create = []
            to_update = []
            for r in records:
                key = f"{r['employee_id']}|{r['work_date']}"
                fields = {
                    "Date": r["work_date"],
                    "ID": r["employee_id"],
                    "Full Name": r["employee_name"],
                    "Department": r["department"] or "--",
                    "Clock In": r["time_in"],
                    "Clock Out": r["time_out"],
                    "Status": r["status"].capitalize(),
                }
                if key in existing_map:
                    to_update.append({"record_id": existing_map[key], "fields": fields})
                else:
                    to_update.append({"fields": fields})

            create_batch = [r for r in to_update if "record_id" not in r]
            update_batch = [r for r in to_update if "record_id" in r]

            for i in range(0, len(create_batch), 100):
                batch = create_batch[i:i+100]
                resp = req.post(f"{base_url}/records/batch_create", headers=headers, json={"records": batch})
                result = resp.json()
                if result.get("code") == 0:
                    pushed += len(result.get("data", {}).get("records", []))

            for i in range(0, len(update_batch), 100):
                batch = update_batch[i:i+100]
                resp = req.post(f"{base_url}/records/batch_update", headers=headers, json={"records": batch})
                result = resp.json()
                if result.get("code") == 0:
                    pushed += len(result.get("data", {}).get("records", []))
        else:
            for i in range(0, len(records), 100):
                batch = records[i:i+100]
                bitable_records = [{
                    "fields": {
                        "Date": r["work_date"],
                        "ID": r["employee_id"],
                        "Full Name": r["employee_name"],
                        "Department": r["department"] or "--",
                        "Clock In": r["time_in"],
                        "Clock Out": r["time_out"],
                        "Status": r["status"].capitalize(),
                    }
                } for r in batch]
                resp = req.post(f"{base_url}/records/batch_create", headers=headers, json={"records": bitable_records})
                result = resp.json()
                if result.get("code") != 0:
                    print(f"[Lark Base Push Error] code={result.get('code')} msg={result.get('msg')}")
                else:
                    pushed += len(result.get("data", {}).get("records", []))

        if append_only or upsert:
            return len(records), pushed
    except Exception as e:
        print(f"[Lark Base Push Error] {e}")

@app.template_filter("day_name")
def day_name_filter(date_str):
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime("%A")
    except Exception:
        return ""


@app.route("/")
def index():
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    full_name = request.args.get("full_name", "")
    department = request.args.get("department", "")
    employee_type = request.args.get("employee_type", "")
    page = int(request.args.get("page", 1))
    per_page = 20

    today = datetime.now().strftime("%Y-%m-%d")
    if not date_from:
        date_from = datetime(datetime.now().year, 5, 1).strftime("%Y-%m-%d")
    if not date_to:
        date_to = today

    records = query_records(
        date_from=date_from or None,
        date_to=date_to or None,
        full_name=full_name or None,
        department=department or None,
        employee_type=employee_type or None,
    )

    employees = query_employees()
    if employees:
        try:
            date_from_int = int(datetime.strptime(date_from, "%Y-%m-%d").strftime("%Y%m%d"))
            date_to_int = int(datetime.strptime(date_to, "%Y-%m-%d").strftime("%Y%m%d"))
            dept_map = _build_dept_map()
            employee_map = {e["user_id"]: e for e in employees}
            user_ids = [e["user_id"] for e in employees]

            all_results = []
            for uid in user_ids:
                try:
                    current_date = datetime.strptime(date_from, "%Y-%m-%d")
                    end_date = datetime.strptime(date_to, "%Y-%m-%d")
                    while current_date <= end_date:
                        batch_end_date = min(current_date + timedelta(days=29), end_date)
                        batch_from = int(current_date.strftime("%Y%m%d"))
                        batch_to = int(batch_end_date.strftime("%Y%m%d"))
                        result = lark.query_attendance(
                            user_ids=[uid],
                            date_from=batch_from,
                            date_to=batch_to,
                            employee_type="employee_id",
                        )
                        items = result.get("data", {}).get("user_task_results", [])
                        all_results.extend(items)
                        current_date = batch_end_date + timedelta(days=1)
                except Exception:
                    continue

            auto_records = []
            for item in all_results:
                user_id = item.get("user_id", "")
                emp = employee_map.get(user_id, {})
                day = item.get("day", "")
                work_date = f"{str(day)[:4]}-{str(day)[4:6]}-{str(day)[6:8]}" if day else ""
                employee_name = item.get("employee_name", "") or emp.get("name", user_id)

                time_in = "--"
                time_out = "--"
                status = "absent"
                photo_in = ""
                photo_out = ""

                for record in item.get("records", []):
                    check_in_result = record.get("check_in_result", "")
                    check_out_result = record.get("check_out_result", "")
                    check_in_record = record.get("check_in_record") or {}
                    check_out_record = record.get("check_out_record") or {}

                    if check_in_record.get("check_time"):
                        ts = int(check_in_record["check_time"])
                        time_in = datetime.fromtimestamp(ts).strftime("%I:%M %p")
                        photos = check_in_record.get("photo_urls", [])
                        if photos:
                            photo_in = photos[0]
                    if check_out_record.get("check_time"):
                        ts = int(check_out_record["check_time"])
                        time_out = datetime.fromtimestamp(ts).strftime("%I:%M %p")
                        photos = check_out_record.get("photo_urls", [])
                        if photos:
                            photo_out = photos[0]

                    if check_in_result == "NoNeedCheck" and check_out_result == "NoNeedCheck":
                        supplement = record.get("check_in_result_supplement", "")
                        status = "leave" if supplement == "Leave" else "rest day"
                    elif time_in != "--" and time_out != "--":
                        status = "present"
                    elif time_in != "--":
                        status = "present"
                    if check_in_result == "Late":
                        status = "late"

                dept_id = emp.get("department_id", "")
                department_name = emp.get("department", "")
                if not department_name and dept_id and dept_id in dept_map:
                    department_name = dept_map[dept_id]

                auto_records.append((
                    user_id, employee_name, department_name,
                    emp.get("employee_type", "staff"), work_date,
                    time_in, time_out, status, photo_in, photo_out,
                ))

            if auto_records:
                clear_records_by_date(date_from, date_to)
                insert_records(auto_records)

                conn = get_db()
                cur = conn.cursor()
                for ar in auto_records:
                    uid, name, dept, emp_type = ar[0], ar[1], ar[2], ar[3]
                    if name:
                        cur.execute("UPDATE employees SET name = %s WHERE user_id = %s AND (name = '' OR name IS NULL)", (name, uid))
                    if dept:
                        cur.execute("UPDATE employees SET department = %s WHERE user_id = %s AND (department = '' OR department IS NULL)", (dept, uid))
                conn.commit()
                conn.close()

                all_db_records = query_records(date_from=None, date_to=None)
                _push_to_lark_base(all_db_records, append_only=True)

            records = query_records(
                date_from=date_from or None,
                date_to=date_to or None,
                full_name=full_name or None,
                department=department or None,
                employee_type=employee_type or None,
            )
            records.sort(key=lambda r: (r["work_date"], r["employee_name"]), reverse=True)
        except Exception:
            pass

    stats = get_stats()

    total_records = len(records)
    total_pages = max(1, (total_records + per_page - 1) // per_page)
    if page < 1:
        page = 1
    if page > total_pages:
        page = total_pages
    start = (page - 1) * per_page
    end = start + per_page
    paginated_records = records[start:end]

    departments = list({r["department"] for r in records if r["department"]})
    departments.sort()

    emp_stats = get_employee_stats()

    return render_template(
        "index.html",
        records=paginated_records,
        all_records=records,
        stats=stats,
        departments=departments,
        employees=employees,
        emp_stats=emp_stats,
        filters={
            "date_from": date_from,
            "date_to": date_to,
            "full_name": full_name,
            "department": department,
            "employee_type": employee_type,
        },
        today=today,
        page=page,
        total_pages=total_pages,
        total_records=total_records,
    )


@app.route("/sync-users", methods=["POST"])
def sync_users():
    try:
        users = lark.get_all_users()

        rows = []
        for u in users:
            employee_id = u.get("user_id", "")
            open_id = u.get("open_id", "")
            name = u.get("name", "")
            en_name = u.get("en_name", "")
            emails = u.get("emails", [])
            email = emails[0] if emails else u.get("email", "")
            mobiles = u.get("mobiles", [])
            mobile = mobiles[0] if mobiles else ""
            dept_ids = u.get("department_ids", [])
            dept_id = dept_ids[0] if dept_ids else ""
            job_title = u.get("job_title", "")
            avatar = u.get("avatar", {}).get("avatar_72", "")
            status = "active" if u.get("status", {}).get("is_activated") else "inactive"

            emp_type = "staff"
            if job_title:
                title_lower = job_title.lower()
                if any(w in title_lower for w in ["worker", "laborer", "operator", "technician", "helper"]):
                    emp_type = "worker"

            rows.append((
                employee_id, name, en_name, email, mobile,
                "", dept_id, job_title, emp_type, avatar, status,
            ))

        if rows:
            clear_employees()
            insert_employees(rows)

            dept_map = _build_dept_map()
            conn = get_db()
            cur = conn.cursor()
            for row in rows:
                dept_id = row[6]
                if dept_id and dept_id in dept_map:
                    cur.execute(
                        "UPDATE employees SET department = %s WHERE department_id = %s",
                        (dept_map[dept_id], dept_id),
                    )
            conn.commit()
            conn.close()

        flash(f"Synced {len(rows)} employees from Lark.", "success")
    except Exception as e:
        flash(f"Sync failed: {str(e)}", "error")

    return redirect(url_for("index"))


def _build_dept_map():
    try:
        depts = lark.get_departments()
        return {d.get("department_id", ""): d.get("name", "") for d in depts}
    except Exception:
        return {}


@app.route("/add-employee", methods=["POST"])
def add_employee():
    try:
        employee_id = request.form.get("employee_id", "").strip()
        department = request.form.get("department", "").strip()

        if not employee_id:
            flash("Employee ID is required.", "error")
            return redirect(url_for("index"))

        existing = query_employees()
        existing_ids = [e["user_id"] for e in existing]
        if employee_id in existing_ids:
            flash(f"Employee {employee_id} already exists.", "error")
            return redirect(url_for("index"))

        insert_employees([(
            employee_id, "", "", "", "",
            department, "", "", "staff", "", "active",
        )])

        flash(f"Added employee {employee_id}.", "success")
    except Exception as e:
        flash(f"Failed to add employee: {str(e)}", "error")

    return redirect(url_for("index"))


@app.route("/delete-employee", methods=["POST"])
def delete_employee():
    try:
        employee_id = request.form.get("employee_id", "").strip()
        if employee_id:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("DELETE FROM employees WHERE user_id = %s", (employee_id,))
            conn.commit()
            conn.close()
            flash(f"Deleted employee {employee_id}.", "success")
    except Exception as e:
        flash(f"Failed to delete: {str(e)}", "error")

    return redirect(url_for("index"))


@app.route("/sync-attendance", methods=["POST"])
def sync_attendance():
    try:
        date_from = request.form.get("date_from", "")
        date_to = request.form.get("date_to", "")

        if not date_from or not date_to:
            flash("Please select a date range.", "error")
            return redirect(url_for("index"))

        date_from_int = int(datetime.strptime(date_from, "%Y-%m-%d").strftime("%Y%m%d"))
        date_to_int = int(datetime.strptime(date_to, "%Y-%m-%d").strftime("%Y%m%d"))

        employees = query_employees()
        if not employees:
            flash("No employees found. Please sync users first.", "error")
            return redirect(url_for("index"))

        dept_map = _build_dept_map()
        employee_map = {e["user_id"]: e for e in employees}

        user_ids = [e["user_id"] for e in employees]

        all_results = []
        for uid in user_ids:
            try:
                result = lark.query_attendance(
                    user_ids=[uid],
                    date_from=date_from_int,
                    date_to=date_to_int,
                    employee_type="employee_id",
                )
                items = result.get("data", {}).get("user_task_results", [])
                all_results.extend(items)
            except Exception:
                continue

        records = []
        for item in all_results:
            user_id = item.get("user_id", "")
            emp = employee_map.get(user_id, {})

            day = item.get("day", "")
            work_date = f"{str(day)[:4]}-{str(day)[4:6]}-{str(day)[6:8]}" if day else ""

            employee_name = item.get("employee_name", "")
            if not employee_name:
                employee_name = emp.get("name", user_id)

            time_in = "--"
            time_out = "--"
            status = "absent"
            photo_in = ""
            photo_out = ""

            for record in item.get("records", []):
                check_in_result = record.get("check_in_result", "")
                check_out_result = record.get("check_out_result", "")

                check_in_record = record.get("check_in_record") or {}
                check_out_record = record.get("check_out_record") or {}

                if check_in_record.get("check_time"):
                    ts = int(check_in_record["check_time"])
                    time_in = datetime.fromtimestamp(ts).strftime("%I:%M %p")
                    photos = check_in_record.get("photo_urls", [])
                    if photos:
                        photo_in = photos[0]

                if check_out_record.get("check_time"):
                    ts = int(check_out_record["check_time"])
                    time_out = datetime.fromtimestamp(ts).strftime("%I:%M %p")
                    photos = check_out_record.get("photo_urls", [])
                    if photos:
                        photo_out = photos[0]

                if check_in_result == "NoNeedCheck" and check_out_result == "NoNeedCheck":
                    supplement = record.get("check_in_result_supplement", "")
                    if supplement == "Leave":
                        status = "leave"
                    else:
                        status = "rest day"
                elif time_in != "--" and time_out != "--":
                    status = "present"
                elif time_in != "--":
                    status = "present"

                if check_in_result == "Late":
                    status = "late"

            dept_id = emp.get("department_id", "")
            department = emp.get("department", "")
            if not department and dept_id and dept_id in dept_map:
                department = dept_map[dept_id]

            records.append((
                user_id,
                employee_name,
                department,
                emp.get("employee_type", "staff"),
                work_date,
                time_in,
                time_out,
                status,
                photo_in,
                photo_out,
            ))

        if records:
            clear_records()
            insert_records(records)

            conn = get_db()
            cur = conn.cursor()
            for rec in records:
                uid, name, dept = rec[0], rec[1], rec[2]
                if name:
                    cur.execute("UPDATE employees SET name = %s WHERE user_id = %s AND (name = '' OR name IS NULL)", (name, uid))
                if dept:
                    cur.execute("UPDATE employees SET department = %s WHERE user_id = %s AND (department = '' OR department IS NULL)", (dept, uid))
            conn.commit()
            conn.close()

            flash(f"Synced {len(records)} attendance records from Lark.", "success")
        else:
            flash("No attendance records found for the selected date range.", "error")

    except Exception as e:
        flash(f"Sync failed: {str(e)}", "error")

    return redirect(url_for("index"))


@app.route("/api/employees")
def api_employees():
    employees = query_employees(
        full_name=request.args.get("full_name"),
        department=request.args.get("department"),
        employee_type=request.args.get("employee_type"),
    )
    stats = get_employee_stats()
    return jsonify({"employees": employees, "stats": stats})


@app.route("/export-csv")
def export_csv():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    full_name = request.args.get("full_name", "")
    department = request.args.get("department", "")
    employee_type = request.args.get("employee_type", "")

    records = query_records(
        date_from=date_from or None,
        date_to=date_to or None,
        full_name=full_name or None,
        department=department or None,
        employee_type=employee_type or None,
    )
    records.sort(key=lambda r: (r["work_date"], r["employee_name"]), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Records"

    headers = ["Date", "Full Name", "Department", "Type", "Clock In", "Clock Out", "Status"]
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4facfe", end_color="4facfe", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row_fill_even = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="left", vertical="center")

    status_colors = {
        "present": Font(name="Calibri", size=11, color="059669", bold=True),
        "late": Font(name="Calibri", size=11, color="D97706", bold=True),
        "absent": Font(name="Calibri", size=11, color="DC2626", bold=True),
        "leave": Font(name="Calibri", size=11, color="2563EB", bold=True),
        "rest day": Font(name="Calibri", size=11, color="6B7280", bold=True),
    }

    for row_idx, r in enumerate(records, 2):
        values = [
            r["work_date"],
            r["employee_name"],
            r["department"] or "--",
            r["employee_type"].capitalize(),
            r["time_in"],
            r["time_out"],
            r["status"].capitalize(),
        ]
        fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = status_colors.get(r["status"], data_font) if col_idx == 7 else data_font
            cell.fill = fill
            cell.alignment = data_align
            cell.border = thin_border

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    ws.auto_filter.ref = f"A1:G{len(records) + 1}"
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{date_from}_to_{date_to}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/api/records")
def api_records():
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    status = request.args.get("status")
    records = query_records(
        date_from=date_from or None,
        date_to=date_to or None,
    )
    if status:
        records = [r for r in records if r["status"] == status]
    return jsonify({"records": records})


@app.route("/photo")
def proxy_photo():
    import requests as req
    photo_url = request.args.get("url", "")
    if not photo_url:
        return "", 404
    try:
        resp = req.get(photo_url, headers={"Authorization": f"Bearer {lark._get_tenant_access_token()}"})
        return Response(resp.content, content_type=resp.headers.get("Content-Type", "image/jpeg"))
    except Exception:
        return "", 404


@app.route("/api/attendance/check-face", methods=["POST"])
def check_face_stub():
    return jsonify({"ok": True})

@app.route("/sync-to-lark", methods=["POST"])
def sync_to_lark():
    try:
        records = query_records(date_from=None, date_to=None)

        if not records:
            return jsonify({"ok": False, "message": "No records to sync."}), 400

        total, added = _push_to_lark_base(records, append_only=True)
        return jsonify({"ok": True, "message": f"{total} records checked, {added} new records added to Lark Base.", "total": total, "added": added})
    except Exception as e:
        return jsonify({"ok": False, "message": f"Sync failed: {str(e)}"}), 500


AUTO_SYNC_INTERVAL = 1800  # seconds (30 minutes)


def _auto_sync():
    """Background thread: fetch today's attendance from Lark API, update local DB only."""
    try:
        employees = query_employees()
        if not employees:
            return

        today = datetime.now().strftime("%Y-%m-%d")
        date_from = int(datetime.strptime(today, "%Y-%m-%d").strftime("%Y%m%d"))
        date_to = date_from

        dept_map = _build_dept_map()
        employee_map = {e["user_id"]: e for e in employees}
        user_ids = [e["user_id"] for e in employees]

        all_results = []
        for uid in user_ids:
            try:
                result = lark.query_attendance(
                    user_ids=[uid],
                    date_from=date_from,
                    date_to=date_to,
                    employee_type="employee_id",
                )
                items = result.get("data", {}).get("user_task_results", [])
                all_results.extend(items)
            except Exception:
                continue

        records = []
        for item in all_results:
            user_id = item.get("user_id", "")
            emp = employee_map.get(user_id, {})
            day = item.get("day", "")
            work_date = f"{str(day)[:4]}-{str(day)[4:6]}-{str(day)[6:8]}" if day else ""
            employee_name = item.get("employee_name", "") or emp.get("name", user_id)

            time_in = "--"
            time_out = "--"
            status = "absent"
            photo_in = ""
            photo_out = ""

            for record in item.get("records", []):
                check_in_result = record.get("check_in_result", "")
                check_out_result = record.get("check_out_result", "")
                check_in_record = record.get("check_in_record") or {}
                check_out_record = record.get("check_out_record") or {}

                if check_in_record.get("check_time"):
                    ts = int(check_in_record["check_time"])
                    time_in = datetime.fromtimestamp(ts).strftime("%I:%M %p")
                    photos = check_in_record.get("photo_urls", [])
                    if photos:
                        photo_in = photos[0]
                if check_out_record.get("check_time"):
                    ts = int(check_out_record["check_time"])
                    time_out = datetime.fromtimestamp(ts).strftime("%I:%M %p")
                    photos = check_out_record.get("photo_urls", [])
                    if photos:
                        photo_out = photos[0]

                if check_in_result == "NoNeedCheck" and check_out_result == "NoNeedCheck":
                    supplement = record.get("check_in_result_supplement", "")
                    status = "leave" if supplement == "Leave" else "rest day"
                elif time_in != "--" and time_out != "--":
                    status = "present"
                elif time_in != "--":
                    status = "present"
                if check_in_result == "Late":
                    status = "late"

            dept_id = emp.get("department_id", "")
            department_name = emp.get("department", "")
            if not department_name and dept_id and dept_id in dept_map:
                department_name = dept_map[dept_id]

            records.append((
                user_id, employee_name, department_name,
                emp.get("employee_type", "staff"), work_date,
                time_in, time_out, status, photo_in, photo_out,
            ))

        if records:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("DELETE FROM attendance_records WHERE work_date = %s", (today,))
            conn.commit()
            conn.close()
            insert_records(records)

            conn = get_db()
            cur = conn.cursor()
            for rec in records:
                uid, name, dept = rec[0], rec[1], rec[2]
                if name:
                    cur.execute("UPDATE employees SET name = %s WHERE user_id = %s AND (name = '' OR name IS NULL)", (name, uid))
                if dept:
                    cur.execute("UPDATE employees SET department = %s WHERE user_id = %s AND (department = '' OR department IS NULL)", (dept, uid))
            conn.commit()
            conn.close()

            today_records = query_records(date_from=today, date_to=today)
            _push_to_lark_base(today_records, upsert=True)
    except Exception as e:
        print(f"[Auto-Sync Error] {e}")


def _auto_sync_loop():
    while True:
        threading.Thread(target=_auto_sync, daemon=True).start()
        threading.Event().wait(AUTO_SYNC_INTERVAL)


if __name__ == "__main__":
    sync_thread = threading.Thread(target=_auto_sync_loop, daemon=True)
    sync_thread.start()
    app.run(debug=True)
